import os
import csv
import sys
import time
import shutil
import subprocess
import threading
import Queue
import string
import logging
from time import strftime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment


# Don't forget to:
# 1. Create new <test class> and in case of new testclesses in .jar. And add it to <testList>
# 2. Create new <browser class> in case of new browsers to test. And add it to <broList>

voltage = "3.7"


def RunMonitor(measurementDuration):
    os.chdir("C:/Program Files (x86)/Monsoon Solutions Inc/Power Monitor")
    print "start measurement at " + strftime("%m-%d %H:%M:%S")
    logging.debug( u"start measurement")
    os.system(
        "PowerToolCmd.exe -trigger=ETY100D" + measurementDuration + "A -vout=" + voltage + " -USB=auto -keeppower -savefile=battery_test.pt4 -noexitwait")
    os.chdir(homeDir)
    shutil.copy("C:/Program Files (x86)/Monsoon Solutions Inc/Power Monitor/battery_test.csv", "battery_test.csv")
    print "stop measurement at " + strftime("%m-%d %H:%M:%S")
    logging.debug( u"stop measurement")


def getNames(*arg):
    class AsynchronousFileReader(threading.Thread):
        def __init__(self, fd, queue):
            assert isinstance(queue, Queue.Queue)
            assert callable(fd.readline)
            threading.Thread.__init__(self)
            self._fd = fd
            self._queue = queue

        def run(self):
            for line in iter(self._fd.readline, ''):
                self._queue.put(line)

        def eof(self):
            return not self.is_alive() and self._queue.empty()

        def stop(self):
            self._stop.set()

    if (arg[0] == "browser"):
        commandString = ["adb", "shell", "dumpsys", "package", arg[1], "|", "grep", "versionName"]
        numberInOut = 1
    elif (arg[0] == "device"):
        commandString = ["adb", "shell", "getprop", "ro.product.model"]
        numberInOut = 0

    process = subprocess.Popen(commandString,
                                   stdout=subprocess.PIPE)
    stdout_queue = Queue.Queue()
    stdout_reader = AsynchronousFileReader(process.stdout, stdout_queue)
    stdout_reader.start()
    yaBroVersion = stdout_queue.get().split("=")[numberInOut].strip()
    return yaBroVersion


def LogReader(measurementDuration):
    class AsynchronousFileReader(threading.Thread):

        def __init__(self, fd, queue):
            assert isinstance(queue, Queue.Queue)
            assert callable(fd.readline)
            threading.Thread.__init__(self)
            self._fd = fd
            self._queue = queue

        def run(self):
            for line in iter(self._fd.readline, ''):
                self._queue.put(line)

        def eof(self):
            return not self.is_alive() and self._queue.empty()

        def stop(self):
            self._stop.set()

    process = subprocess.Popen(["adb", "logcat", "-v", "time"], stdout=subprocess.PIPE)

    stdout_queue = Queue.Queue()
    stdout_reader = AsynchronousFileReader(process.stdout, stdout_queue)
    stdout_reader.start()

    while not stdout_reader.eof():
        while not stdout_queue.empty():
            line = stdout_queue.get()
            if "start measurement" in line:
                RunMonitor(measurementDuration)


def LogFailFinder():
    class AsynchronousFileReader(threading.Thread):

        def __init__(self, fd, queue):
            assert isinstance(queue, Queue.Queue)
            assert callable(fd.readline)
            threading.Thread.__init__(self)
            self._fd = fd
            self._queue = queue

        def run(self):
            for line in iter(self._fd.readline, ''):
                self._queue.put(line)

        def eof(self):
            return not self.is_alive() and self._queue.empty()

        def stop(self):
            self._stop.set()

    process = subprocess.Popen(["adb", "logcat", "-v", "time", "-d"], stdout=subprocess.PIPE)

    stdout_queue = Queue.Queue()
    stdout_reader = AsynchronousFileReader(process.stdout, stdout_queue)
    stdout_reader.start()

    while not stdout_reader.eof():
        while not stdout_queue.empty():
            line = stdout_queue.get()
            if "battery test failed" in line:
                return line
            else:
                return "passed"


def RunTests(broList, browser, testList, test):
    if (len(sys.argv)) == 3:
        print "control test"
        print "rebooting device..."
        logging.debug( u"rebooting device...")
        os.system("adb reboot")
        time.sleep(60)
        os.system(
            "adb shell uiautomator runtest /data/local/tmp/battery-test.jar -c ru.batterytest.UnlockDevice")
        print "...please wait..."
        logging.debug( u"...please wait...")
        time.sleep(300)
        print "control test"
        logging.debug( u"control test")
        os.system(
            "start adb shell uiautomator runtest /data/local/tmp/battery-test.jar -c ru.batterytest.Control --nohup")
        LogReader(str(540))
        time.sleep(3)
        try:
            with open('battery_test.csv') as csvfile:
                testResults = csv.DictReader(csvfile)
                currentList = []

                for row in testResults:
                    current = row['Main Avg Power (mW)']
                    currentList.append(current)

                currentList = map(int, currentList)
                try:
                    print "\n" + strftime(
                        "%m-%d %H:%M:%S") + " Control Test: " + str(sum(currentList) / len(currentList)) + "\n"
                    logging.debug( u"Control Test: : " + str(sum(currentList) / len(currentList)))
                except:
                    print "Control Test result printing error"
                    logging.debug( u"Control Test result printing error")
        except:
            print "battery_test.csv reading error\n"
            logging.debug( u"battery_test.csv reading error")

        try:
            worksheet["A1"] = str(sum(currentList) / len(currentList))
            workbook.save(xlsxFilename)
        except:
            print "Control Test result writing error\n"
            logging.debug( u"Control Test result writing error")
        time.sleep(5)
    elif "--no-control" in sys.argv[3]:
        print "skipping Control Test"
        logging.debug( u"skipping Control Test")

    for browserToRun in browser:
        print browserToRun.browserName  + " tests started"
        logging.debug( u"" + browserToRun.browserName  + " tests started")

        for testToRun in test:
            print browserToRun.browserName + " " + testToRun.testClass  + " started"
            logging.debug( u"" + browserToRun.browserName + " " + testToRun.testClass  + " started")
            print "rebooting device..."
            logging.debug( u"rebooting device...")
            os.system("adb reboot")
            time.sleep(60)
            os.system(
                "adb shell uiautomator runtest /data/local/tmp/battery-test.jar -c ru.batterytest.UnlockDevice")
            print "...please wait..."
            logging.debug( u"...please wait...")
            time.sleep(300)

            testNumber = 1
            allTestsCurrentAvg = []
            print testToRun.testClass + " started"
            logging.debug( u""+testToRun.testClass + " started")

            for browsersToClear in broList:
                print browsersToClear.browserName + " clearing..."
                logging.debug( u"" + browsersToClear.browserName + " clearing...")
                os.system("adb shell pm clear " + browsersToClear.package)
                time.sleep(1)

            if hasattr(testToRun, 'notFirstStart'):
                print "First start..."
                logging.debug( u"First start...")
                os.system(
                    "adb shell uiautomator runtest /data/local/tmp/battery-test.jar -c ru.batterytest.steps.BrowserFirstStart" + " -e browser " + browserToRun.testBrowser)
                print "...please wait..."
                logging.debug( u"...please wait...")
                time.sleep(80)

            if hasattr(testToRun, 'enableRotation'):
                os.system(
                    "adb shell content insert --uri content://settings/system --bind name:s:accelerometer_rotation --bind value:i:1")
                print "Screen rotation enabled!"
                logging.debug( u"Screen rotation enabled!")

            retryCount = 0
            while testNumber < testToRun.runs + 1:

                os.system("adb shell am force-stop " + browserToRun.package)
                print browserToRun.browserName + " stopped!"
                logging.debug( u"" + browserToRun.browserName + " stopped!")

                if hasattr(testToRun, 'clearBrowser'):
                    print browserToRun.browserName + " clearing..."
                    logging.debug( u"" + browserToRun.browserName + " clearing...")
                    os.system("adb shell pm clear " + browserToRun.package)

                # os.system("adb kill-server")
                os.system("adb devices")
                os.system("adb logcat -c")
                os.system(
                    "start adb shell uiautomator runtest /data/local/tmp/battery-test.jar -c ru.batterytest." + browserToRun.testBrowser + "." + testToRun.testClass + " -e browser " + browserToRun.testBrowser + " --nohup")
                LogReader(str(testToRun.measurementDuration))
                time.sleep(3)
                findFailInLog = LogFailFinder()
                if "failed" in findFailInLog:
                    retryCount = retryCount + 1
                    try:
                        print strftime(
                            "%m-%d %H:%M:%S") + " " + browserToRun.browserName + " " + testToRun.testClass + " " + str(
                            testNumber) + ": failed! Count = " + str(
                            retryCount) + "\n" + findFailInLog.strip() + "\n"
                        logging.debug( u"" + browserToRun.browserName + " " + testToRun.testClass + " " + str(testNumber) + ": failed! Count = " + str(retryCount) + "\n" + findFailInLog.strip())
                    except:
                        print "fail result printing error\n"
                        logging.debug( u"fail result printing error")
                    if retryCount == 2:
                        testNumber = testNumber + 1
                        retryCount = 0
                        continue
                else:
                    try:
                        with open('battery_test.csv') as csvfile:
                            testResults = csv.DictReader(csvfile)
                            currentList = []

                            for row in testResults:
                                current = row['Main Avg Power (mW)']
                                currentList.append(current)

                            currentList = map(int, currentList)
                            try:
                                print "\n" + strftime(
                                    "%m-%d %H:%M:%S") + " " + browserToRun.browserName + " " + testToRun.testClass + " " + str(
                                    testNumber) + ": " + str(sum(currentList) / len(currentList)) + "\n"
                                logging.debug( u"" + browserToRun.browserName + " " + testToRun.testClass + " " + str(
                                    testNumber) + ": " + str(sum(currentList) / len(currentList)))
                            except:
                                print "single result printing error"
                                logging.debug( u"single result printing error")

                    except:
                        print "battery_test.csv reading error\n"
                        logging.debug( u"battery_test.csv reading error")

                    testDir = allTestsDir + '/' + browserToRun.browserName + '/' + testToRun.testClass
                    if not os.path.isdir(testDir):
                        os.makedirs(testDir)
                    shutil.copyfile('battery_test.csv', allTestsDir + "/" + browserToRun.browserName + "/" + testToRun.testClass + "/" + testToRun.testClass + "_" + str(testNumber) + "_data.csv")
                    os.system('adb logcat -d > ' + '"' + allTestsDir + '/' + browserToRun.browserName + '/' + testToRun.testClass + '/' + testToRun.testClass + '_' + str(testNumber) + '_log.txt"')

                    testNumber = testNumber + 1
                    allTestsCurrentAvg.append(sum(currentList) / len(currentList))

                time.sleep(5)

            if testToRun.runs > 9 and testNumber > 5:
                for i in list(range(int(float(testToRun.runs)/10.0*2.5))):
                    allTestsCurrentAvg.remove(max(allTestsCurrentAvg))
                    allTestsCurrentAvg.remove(min(allTestsCurrentAvg))

            if hasattr(testToRun, 'enableRotation'):
                os.system(
                    "adb shell content insert --uri content://settings/system --bind name:s:accelerometer_rotation --bind value:i:0")
                print "Screen rotation disabled!"
                logging.debug( u"Screen rotation disabled!")

            try:
                worksheet[str(list(string.ascii_uppercase)[testList.index(testToRun) + 1]) + str(
                    broList.index(browserToRun) + 2)] = sum(allTestsCurrentAvg) / len(allTestsCurrentAvg)
                workbook.save(xlsxFilename)
            except:
                print "full test result writing error\n"
                logging.debug( u"full test result writing error")
            try:
                print strftime(
                    "%m-%d %H:%M:%S") + " " + browserToRun.browserName + " " + testToRun.testClass + " Current Avg: " + str(
                    sum(allTestsCurrentAvg) / len(allTestsCurrentAvg))
                logging.debug( u"" + browserToRun.browserName + " " + testToRun.testClass + " Current Avg: " + str(
                    sum(allTestsCurrentAvg) / len(allTestsCurrentAvg)))
            except:
                print "full test result printing error\n"
                logging.debug( u"full test result printing error")
            time.sleep(5)
            os.system('adb bugreport > ' + '"' + allTestsDir + '/' + browserToRun.browserName + '/' + testToRun.testClass + '/' + testToRun.testClass + '_bugreport.txt"')
            print browserToRun.browserName + " " + testToRun.testClass  + " finished"
            logging.debug( u"" + browserToRun.browserName + " " + testToRun.testClass  + " finished")
        print browserToRun.browserName  + " tests finished"
        logging.debug( u"" + browserToRun.browserName  + " tests finished")

bro = []
test = []


class YandexBrowser:
    package = "com.yandex.browser"
    browserName = "Yandex Browser " + getNames("browser", package)
    testBrowser = "yabro"
    forArgs = "Y"


class Chrome:
    package = "com.android.chrome"
    browserName = "Chrome " + getNames("browser", package)
    testBrowser = "chrome"
    forArgs = "C"


#class Opera:
#    package = "com.opera.browser"
#    browserName = "Opera " + getNames("browser", package)
#    testBrowser = "opera"
#    forArgs = "O"


class ColdStart:
    testClass = "ColdStart"
    measurementDuration = 30
    runs = 20
    notFirstStart = ""
    forArgs = "Cs"


class Foreground:
    testClass = "Foreground"
    measurementDuration = 500
    runs = 1
    clearBrowser = ""
    forArgs = "Fg"


class Background:
    testClass = "Background"
    measurementDuration = 500
    runs = 1
    clearBrowser = ""
    forArgs = "Bg"


class UlrOpen:
    testClass = "UrlOpen"
    measurementDuration = 30
    runs = 20
    clearBrowser = ""
    forArgs = "Uo"


class TenSitesForeground:
    testClass = "TenSitesForeground"
    measurementDuration = 500
    runs = 1
    clearBrowser = ""
    forArgs = "Tsf"


class TenSitesBackground:
    testClass = "TenSitesBackground"
    measurementDuration = 500
    runs = 1
    clearBrowser = ""
    forArgs = "Tsb"


class VideoPlay:
    testClass = "VideoPlay"
    measurementDuration = 500
    runs = 1
    clearBrowser = ""
    enableRotation = ""
    forArgs = "Vp"


class Scroll:
    testClass = "Scroll"
    measurementDuration = 550
    runs = 1
    clearBrowser = ""
    forArgs = "Sc"


class MusicPlay:
    testClass = "MusicPlay"
    measurementDuration = 550
    runs = 1
    clearBrowser = ""
    forArgs = "Mp"


class HundredSitesForeground:
    testClass = "HundredSitesForeground"
    measurementDuration = 500
    runs = 1
    clearBrowser = ""
    forArgs = "Hsf"

class MetrikaLoggingOn:
    testClass = "MetrikaLoggingOn"
    measurementDuration = 170
    runs = 1
    clearBrowser = ""
    forArgs = "MOn"

class MetrikaLoggingOff:
    testClass = "MetrikaLoggingOff"
    measurementDuration = 170
    runs = 1
    clearBrowser = ""
    forArgs = "MOf"


broList = [YandexBrowser, Chrome]
testList = [ColdStart, Foreground, Background, UlrOpen, TenSitesForeground, TenSitesBackground, VideoPlay, Scroll,
            MusicPlay, HundredSitesForeground]

if (len(sys.argv)) > 2:
    for i in broList:
        if i.forArgs in sys.argv[1]:
            bro.append(i)
    for i in testList:
        if i.forArgs in sys.argv[2]:
            test.append(i)
else:
    print "Usage:\n  uiautomator.py <browsers> <tests>\nFor example:\n  asd.py YCO CsFgTsb --no-control\n"
    print "Browser list:"
    for browser in broList:
        print browser.forArgs + " - " + browser.browserName
    print "\nTest list:"
    for n in testList:
        print n.forArgs + " - " + n.testClass
    print "--no-control - for no control measurement"
    sys.exit()


logging.basicConfig(format = u'%(levelname)-8s [%(asctime)s] %(message)s', level = logging.DEBUG, filename = u'uiautomator.log')

allTestsDir = getNames("device") + " " + YandexBrowser.browserName
if not os.path.isdir(allTestsDir):
    os.mkdir(allTestsDir)
homeDir =  os.getcwd()
xlsxFilename = homeDir + "/" + allTestsDir + "/" + "result_table.xlsx"

if not os.path.isfile(xlsxFilename):
    workbookTmp = Workbook()
    workbookTmp.save(xlsxFilename)
workbook = load_workbook(xlsxFilename)
worksheet = workbook.active
alignment = Alignment(horizontal='right')

bNumber = 1
for browserToChoose in broList:
    worksheet["A" + str(broList.index(browserToChoose) + 2)] = browserToChoose.browserName
    logging.debug( u"" + browserToChoose.browserName)
    worksheet.column_dimensions["A"].width = max(len(x.browserName) for x in broList)
    worksheet["A" + str(broList.index(browserToChoose) + 2)].alignment = alignment
    bNumber = bNumber + 1
workbook.save(xlsxFilename)

print ""
columnName = dict()

tNumber = 1
for testToChoose in testList:
    cellNumber = str(list(string.ascii_uppercase)[tNumber]) + "1"
    worksheet[cellNumber] = testToChoose.testClass
    logging.debug( u"" + testToChoose.testClass)
    worksheet[cellNumber].alignment = alignment
    worksheet.column_dimensions[str(list(string.ascii_uppercase)[tNumber])].width = len(testToChoose.testClass) + 2
    tNumber = tNumber + 1
workbook.save(xlsxFilename)

# Uncomment block below to rebuild .jar
#if os.path.isfile("build.xml"):
#    os.remove("build.xml")
#if os.path.isdir("bin"):
#    shutil.rmtree("bin")
#os.system("android.bat create uitest-project -n battery-test -t 2 -p " + homeDir)
#os.system("ant build")

os.system("adb push bin/battery-test.jar /data/local/tmp/")

RunTests(broList, bro, testList, test)
